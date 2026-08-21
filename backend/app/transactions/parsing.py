"""CSV/XLSX parsing + column mapping (task 3.2, FR-2.1/FR-2.2).

Deliberately DB-free so it can be unit-tested in isolation: it turns raw file
bytes into `ParsedRow`s (date, description, signed amount, optional category
name, optional explicit type) plus a list of human-readable per-row errors. The
service layer resolves category ids and the final income/expense `type` against
the database.

Column headers are matched by normalised aliases, so real-world exports
("Transaction Date", "Amount (INR)", "Narration", …) map without the user
hand-editing them. Rows that can't yield a date + amount are skipped and
reported rather than aborting the whole file (NFR-6).
"""

from __future__ import annotations

import csv
import datetime as dt
import io
from dataclasses import dataclass
from decimal import Decimal, InvalidOperation


class UploadParseError(Exception):
    """Fatal, file-level problem (bad type / unreadable / no usable columns)."""


@dataclass
class ParsedRow:
    date: dt.date
    amount: Decimal  # signed as it appeared in the file
    description: str | None
    category_name: str | None
    explicit_type: str | None  # "income" | "expense" | None
    source_row: int  # 1-based line number in the file (for user-facing messages)


@dataclass
class ParsedUpload:
    rows: list[ParsedRow]
    errors: list[str]


# Header alias sets (compared after normalisation: lowercased, alphanumerics only).
_FIELD_ALIASES: dict[str, set[str]] = {
    "date": {"date", "transactiondate", "txndate", "posteddate", "valuedate"},
    "amount": {"amount", "amt", "value", "amountinr", "inr", "amountrs"},
    "description": {
        "description", "desc", "details", "narration", "particulars", "memo",
        "note", "notes",
    },
    "category": {"category", "categoryname", "head", "account"},
    "type": {"type", "direction", "incomeexpense", "drcr", "debitcredit"},
}

_INCOME_WORDS = {"income", "credit", "cr", "in", "revenue", "inflow"}
_EXPENSE_WORDS = {"expense", "debit", "dr", "out", "outflow", "cost"}

_DATE_FORMATS = (
    "%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%m/%d/%Y", "%d/%m/%y",
    "%d %b %Y", "%d %B %Y", "%Y/%m/%d",
)


def _normalise(header: str) -> str:
    return "".join(ch for ch in header.lower() if ch.isalnum())


def _build_column_map(headers: list[str]) -> dict[str, int]:
    """Map standard field -> column index, using the first header that matches."""
    mapping: dict[str, int] = {}
    for idx, header in enumerate(headers):
        norm = _normalise(header or "")
        for field, aliases in _FIELD_ALIASES.items():
            if field not in mapping and norm in aliases:
                mapping[field] = idx
    return mapping


def _parse_date(raw: str) -> dt.date | None:
    raw = raw.strip()
    if not raw:
        return None
    for fmt in _DATE_FORMATS:
        try:
            return dt.datetime.strptime(raw, fmt).date()
        except ValueError:
            continue
    return None


def _parse_amount(raw: str) -> Decimal | None:
    # Strip currency symbols/labels, thousands separators and spaces.
    cleaned = (
        raw.replace("₹", "")  # ₹
        .replace("Rs.", "").replace("Rs", "").replace("INR", "").replace("inr", "")
        .replace(",", "").replace(" ", "").strip()
    )
    # Parenthesised numbers are a common accounting notation for negatives.
    if cleaned.startswith("(") and cleaned.endswith(")"):
        cleaned = "-" + cleaned[1:-1]
    if not cleaned:
        return None
    try:
        return Decimal(cleaned)
    except InvalidOperation:
        return None


def _normalise_type(raw: str) -> str | None:
    norm = _normalise(raw)
    if norm in _INCOME_WORDS:
        return "income"
    if norm in _EXPENSE_WORDS:
        return "expense"
    return None


def _cell(row: list[str], idx: int | None) -> str:
    if idx is None or idx >= len(row):
        return ""
    value = row[idx]
    return "" if value is None else str(value)


def _rows_from_csv(content: bytes) -> list[list[str]]:
    text = content.decode("utf-8-sig", errors="replace")
    return [row for row in csv.reader(io.StringIO(text))]


def _rows_from_xlsx(content: bytes) -> list[list[str]]:
    from openpyxl import load_workbook

    try:
        wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    except Exception as exc:  # openpyxl raises various errors on bad files
        raise UploadParseError("Could not read the Excel file.") from exc
    ws = wb.active
    rows: list[list[str]] = []
    for row in ws.iter_rows(values_only=True):
        rows.append([_cell_to_str(c) for c in row])
    wb.close()
    return rows


def _cell_to_str(cell: object) -> str:
    """Stringify an openpyxl cell. Excel date cells come back as datetime/date —
    render them as ISO dates so the date parser (which takes strings) can read
    them instead of choking on a trailing '00:00:00'."""
    if cell is None:
        return ""
    if isinstance(cell, dt.datetime):
        return cell.date().isoformat()
    if isinstance(cell, dt.date):
        return cell.isoformat()
    return str(cell)


def parse_upload(filename: str, content: bytes) -> ParsedUpload:
    """Parse a CSV or XLSX file into rows + per-row errors.

    Raises UploadParseError for file-level problems (unsupported type, empty
    file, or no recognisable date/amount columns).
    """
    lower = filename.lower()
    if lower.endswith(".csv"):
        raw_rows = _rows_from_csv(content)
    elif lower.endswith(".xlsx"):
        raw_rows = _rows_from_xlsx(content)
    else:
        raise UploadParseError("Unsupported file type — upload a .csv or .xlsx file.")

    # Drop fully-empty rows (trailing blank lines etc.).
    raw_rows = [r for r in raw_rows if any((c or "").strip() for c in r)]
    if not raw_rows:
        raise UploadParseError("The file is empty.")

    headers, *data_rows = raw_rows
    columns = _build_column_map(headers)
    if "date" not in columns or "amount" not in columns:
        raise UploadParseError(
            "Couldn't find the required columns — the file needs a date column "
            "and an amount column."
        )

    rows: list[ParsedRow] = []
    errors: list[str] = []
    for i, raw in enumerate(data_rows, start=2):  # row 1 is the header
        date = _parse_date(_cell(raw, columns.get("date")))
        if date is None:
            errors.append(f"Row {i}: missing or unrecognised date — skipped.")
            continue
        amount = _parse_amount(_cell(raw, columns.get("amount")))
        if amount is None:
            errors.append(f"Row {i}: missing or non-numeric amount — skipped.")
            continue

        description = _cell(raw, columns.get("description")).strip() or None
        category_name = _cell(raw, columns.get("category")).strip() or None
        explicit_type = (
            _normalise_type(_cell(raw, columns.get("type")))
            if "type" in columns
            else None
        )
        rows.append(
            ParsedRow(
                date=date,
                amount=amount,
                description=description,
                category_name=category_name,
                explicit_type=explicit_type,
                source_row=i,
            )
        )

    return ParsedUpload(rows=rows, errors=errors)
